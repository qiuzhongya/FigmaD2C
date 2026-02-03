
import cv2
import numpy as np
from PIL import Image
import math

def calculate_ssim(img1_np, img2_np):
    """
    Calculate Structural Similarity Index (SSIM).
    
    Explanation:
    SSIM is a perceptual metric that quantifies image quality degradation caused by processing such as data compression or by losses in data transmission.
    It measures the similarity between two images based on three comparison measurements: luminance, contrast, and structure.
    
    Interpretation:
    - Range: [-1, 1]
    - 1.0: Identical images.
    - > 0.95: Very high similarity, differences are hard to distinguish by human eye.
    """
    # img1_np, img2_np are grayscale numpy arrays
    
    C1 = 6.5025
    C2 = 58.5225
    
    img1 = img1_np.astype(np.float64)
    img2 = img2_np.astype(np.float64)
    
    kernel = cv2.getGaussianKernel(11, 1.5)
    window = np.outer(kernel, kernel.transpose())
    
    mu1 = cv2.filter2D(img1, -1, window)[5:-5, 5:-5]
    mu2 = cv2.filter2D(img2, -1, window)[5:-5, 5:-5]
    
    mu1_sq = mu1**2
    mu2_sq = mu2**2
    mu1_mu2 = mu1 * mu2
    
    sigma1_sq = cv2.filter2D(img1**2, -1, window)[5:-5, 5:-5] - mu1_sq
    sigma2_sq = cv2.filter2D(img2**2, -1, window)[5:-5, 5:-5] - mu2_sq
    sigma12 = cv2.filter2D(img1 * img2, -1, window)[5:-5, 5:-5] - mu1_mu2
    
    ssim_map = ((2 * mu1_mu2 + C1) * (2 * sigma12 + C2)) / ((mu1_sq + mu2_sq + C1) * (sigma1_sq + sigma2_sq + C2))
    return ssim_map.mean()

def calculate_psnr(img1_np, img2_np):
    """
    Calculate Peak Signal-to-Noise Ratio (PSNR).
    
    Explanation:
    PSNR is a metric used to measure the quality of image compression or restoration.
    It quantifies the maximum possible amount of signal that can be represented by a digital image or video.
    
    Interpretation:
    - Range: [0, ∞]
    - Higher is better.
    - 30 dB: Excellent quality.
    - 20 dB: Good quality.
    - 15 dB: Acceptable quality.
    - Below 10 dB: Poor quality.
    """
    mse = np.mean((img1_np.astype(np.float64) - img2_np.astype(np.float64)) ** 2)
    if mse == 0:
        return float('inf')
    return 20 * math.log10(255.0 / math.sqrt(mse))

def calculate_flow(img1_np, img2_np):
    """
    Calculate Average Optical Flow Magnitude (Pixel Shift).
    
    Explanation:
    This metric uses the Farneback optical flow algorithm to estimate the displacement of pixels between two images.
    It calculates the average magnitude of the flow vectors, representing the average "movement" or shift of visual elements.
    
    Interpretation:
    - Unit: Pixels
    - Lower is better (less movement).
    - 0: No pixel shift (static image).
    - High value + High SSIM: Indicates UI elements are similar but have shifted position (layout regression).
    """
    # Farneback optical flow
    flow = cv2.calcOpticalFlowFarneback(img1_np, img2_np, None, 0.5, 3, 15, 3, 5, 1.2, 0)
    mag, _ = cv2.cartToPolar(flow[..., 0], flow[..., 1])
    return np.mean(mag)

# Test
def get_score(img1_path, img2_path):
    img1 = Image.open(img1_path).convert('L')
    img2 = Image.open(img2_path).convert('L')
    
    # Resize to match
    w = min(img1.width, img2.width)
    h = min(img1.height, img2.height)
    img1 = img1.resize((w, h))
    img2 = img2.resize((w, h))
    
    i1 = np.array(img1)
    i2 = np.array(img2)
    try:
        return calculate_ssim(i1, i2), calculate_psnr(i1, i2), calculate_flow(i1, i2)
    except Exception as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"Error: {e}")

import os
import glob
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter   # 顶部加一行

ROOT_DIR = "/Users/bytedance/sampledata"

# ---------- 替换原来的 main() ----------
def main():
    records = {}          # records[folder][i] = (ssim, psnr, flow)
    thumbs   = {}         # thumbs[folder]['compose'][i]  缩略图路径
    IMAGE_W, IMAGE_H = 230, 500  # 插入 Excel 的缩略图尺寸

    # 1. 遍历所有子目录，计算指标并保存缩略图
    for subdir in sorted(next(os.walk(ROOT_DIR))[1]):
        dir_path = os.path.join(ROOT_DIR, subdir)
        figma_path = os.path.join(dir_path, "figma_screenshot.png")
        if not os.path.exists(figma_path):
            print(f"[WARN] 跳过 {subdir}：找不到 figma_screenshot.png")
            continue

        records[subdir] = {}
        thumbs[subdir]   = {'compose': {}, 'figma': figma_path}

        for i in range(1, 7):
            file_name = f"com.example.myapplication_ResourcesTest_compose[Default]0{i}.png"
            d2c_path  = os.path.join(dir_path, file_name)
            if not os.path.exists(d2c_path):
                print(f"[WARN] 跳过 {subdir} 的 {file_name}")
                continue

            scores = get_score(d2c_path, figma_path)
            records[subdir][i] = scores
            thumbs[subdir]['compose'][i] = d2c_path

    # 2. 写 Excel
    wb  = Workbook()
    ws  = wb.active
    ws.title = "D2C_Report"

    # 2-1 表头
    ws.append(["Folder Name", "01", "02", "03", "04", "05", "06", "figma_image"])
    # 列宽
    for col, w in enumerate([12, 35, 35, 35, 35, 35, 35, 35], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # 2-2 逐文件夹写数据
    start_row = 2
    for folder in sorted(records.keys()):
        compose_paths = thumbs[folder]['compose']
        figma_path    = thumbs[folder]['figma']
        # 插图：compose 01–06 → 列 2–7
        for i in range(1, 7):
            if i in compose_paths:
                img = XLImage(compose_paths[i])
                img.width, img.height = IMAGE_W, IMAGE_H
                # 插在 SSIM 那一行
                ws.add_image(img, f"{chr(65+i)}{start_row}")
                content_str = f"ssim: {records[subdir][i][0]}\npsnr: {records[subdir][i][1]}\nflow: {records[subdir][i][2]}\n"
                cell = ws.cell(start_row+1, column=i+1, value=content_str)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 插图：figma → 列 8
        figma_img = XLImage(figma_path)
        figma_img.width, figma_img.height = IMAGE_W, IMAGE_H
        ws.add_image(figma_img, f"H{start_row}")
        ws.cell(start_row, column=1, value=f"{folder}")
        ws.cell(start_row + 1, column=1, value=f"{folder}")

        ws.row_dimensions[start_row].height = IMAGE_H * 0.8
        ws.row_dimensions[start_row + 1].height = 50
        start_row += 2

    # 3. 保存
    out_xlsx = os.path.join(ROOT_DIR, "d2c_report.xlsx")
    wb.save(out_xlsx)
    print(f"报告已生成：{out_xlsx}")
   

if __name__ == "__main__":
    import sys
    main()
