import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

base_dir = "/tmp/d2c_task_output/"

# ---------- 工具函数 ----------
def is_valid_folder(folder_name):
    return not folder_name.isdigit() and any(c.isalpha() for c in folder_name)

def find_log_file(folder_path, folder_name):
    preferred = os.path.join(folder_path, f"{folder_name}.log")
    if os.path.isfile(preferred):
        return preferred
    for f in os.listdir(folder_path):
        if f.endswith(".log"):
            return os.path.join(folder_path, f)
    return None

def calc_duration(log_path):
    with open(log_path, "r", encoding="utf-8") as fh:
        lines = [L for L in fh if L.strip()]
    if len(lines) < 2:
        return 0
    pattern = r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3})"
    first_match = re.search(pattern, lines[0])
    last_match = re.search(pattern, lines[-1])
    if not first_match or not last_match:
        return 0
    fmt = "%Y-%m-%d %H:%M:%S,%f"
    t1 = datetime.strptime(first_match.group(1), fmt)
    t2 = datetime.strptime(last_match.group(1), fmt)
    return (t2 - t1).total_seconds()

def format_duration(sec):
    """≥60s 显示 x分y秒，否则 x秒"""
    if sec < 60:
        return f"{int(sec)}秒"
    m, s = divmod(int(sec), 60)
    return f"{m}分{s}秒"

def find_images(folder_path):
    images_dir = os.path.join(folder_path, "app/src/test/snapshots/images/")
    if not os.path.exists(images_dir):
        return None, None
    com_example_img = None
    figma_screenshot_img = None
    for img_file in os.listdir(images_dir):
        if img_file.startswith("com.example.myapplication"):
            com_example_img = os.path.join(images_dir, img_file)
        elif img_file.startswith("figma_screenshot"):
            figma_screenshot_img = os.path.join(images_dir, img_file)
        if com_example_img and figma_screenshot_img:
            break
    return com_example_img, figma_screenshot_img

# ---------- 主函数 ----------
def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshots"

    # 表头
    headers = ["Folder Name", "D2C Image", "Figma Screenshot", "Log File", "Duration"]
    ws.append(headers)
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = [45, 45, 45, 60, 15][col - 1]

    folders = [f for f in os.listdir(base_dir)
               if os.path.isdir(os.path.join(base_dir, f)) and is_valid_folder(f)]
    total_times = 0
    row_num = 2
    for folder in folders:
        folder_path = os.path.join(base_dir, folder)
        com_example_img, figma_screenshot_img = find_images(folder_path)
        log_file = find_log_file(folder_path, folder)
        duration = calc_duration(log_file) if log_file else 0
        total_times += duration
        if com_example_img and figma_screenshot_img:
            ws.cell(row=row_num, column=1, value=folder)

            # 插入两张图片
            for col_idx, img_path in enumerate([com_example_img, figma_screenshot_img], start=2):
                img = XLImage(img_path)
                img.width, img.height = 117, 252
                ws.add_image(img, f"{chr(64 + col_idx)}{row_num}")

            # 超链接“嵌入”日志文件
            if log_file:
                cell = ws.cell(row=row_num, column=4, value=os.path.basename(log_file))
                cell.hyperlink = log_file
                cell.font = Font(color="0000FF", underline="single")
            else:
                ws.cell(row=row_num, column=4, value="N/A")

            # 格式化时长
            ws.cell(row=row_num, column=5, value=format_duration(duration))

            ws.row_dimensions[row_num].height = 280
            row_num += 1
    average_time = float(total_times) / len(folder)
    out_xlsx = f"d2c_{datetime.now().strftime('%Y%m%d—%H')}_report.xlsx"
    wb.save(out_xlsx)
    print(f"✅ 报告已保存为 {out_xlsx},平均时间：{format_duration(average_time)}")

if __name__ == "__main__":
    main()
