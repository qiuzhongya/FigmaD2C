import os, re, subprocess
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from typing import List
    
base_dir = "/Users/bytedance/task_out01/d2c_task_output0130"
#base_dir = "/Users/bytedance/task_out01/d2c_task_output0130_4"
COMPARE_PY = os.path.join(os.path.dirname(os.path.abspath(__file__)), "image_compare.py")
image_width = 230
image_height = 500
# ---------- 工具函数（保持不变） ----------
def is_valid_folder(name):
    return not name.isdigit() and any(c.isalpha() for c in name)

def find_log_file(fpath, fname):
    preferred = os.path.join(fpath, f"{fname}.log")
    if os.path.isfile(preferred):
        return preferred
    for f in os.listdir(fpath):
        if f.endswith(".log"):
            return os.path.join(fpath, f)
    return None

# ---------- 新增：查找 Greeting.kt 并统计行数 ----------
def find_greeting_lines(fpath: str) -> int:
    """
    在 fpath 下递归查找第一个 Greeting.kt，返回其代码行数（不含空行）。
    找不到返回 0。
    """
    for root, _, files in os.walk(fpath):
        for file in files:
            if file == "Greeting.kt":
                kt_path = os.path.join(root, file)
                with open(kt_path, encoding='utf-8') as fh:
                    # 去掉空行
                    return sum(1 for line in fh if line.strip())
    return 0

def calc_duration(log_path):
    with open(log_path, "r", encoding="utf-8") as fh:
        lines = [L for L in fh if L.strip()]
    if len(lines) < 2:
        return 0
    pattern = r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d{3})"
    first, last = re.search(pattern, lines[0]), re.search(pattern, lines[-1])
    if not (first and last):
        return 0
    fmt = "%Y-%m-%d %H:%M:%S,%f"
    t1 = datetime.strptime(first.group(1), fmt)
    t2 = datetime.strptime(last.group(1), fmt)
    return (t2 - t1).total_seconds()

# ---------- 新增：一次性扫描 token ----------
def calc_tokens(log_path):
    """
    返回 (input_tokens, output_tokens, total_tokens)
    日志里可能有多条，累加即可
    """
    inp = out = total = 0
    if not log_path or not os.path.isfile(log_path):
        return 0, 0, 0
    with open(log_path, encoding="utf-8") as fh:
        for line in fh:
            # 匹配单引号或双引号
            m_in  = re.search(r"['\"]input_tokens['\"]\s*:\s*(\d+)", line)
            m_out = re.search(r"['\"]output_tokens['\"]\s*:\s*(\d+)", line)
            m_tot = re.search(r"['\"]total_tokens['\"]\s*:\s*(\d+)", line)
            if m_in:  inp  += int(m_in.group(1))
            if m_out: out  += int(m_out.group(1))
            if m_tot: total += int(m_tot.group(1))  # 一般只有一行 total
    # 若日志里没给出 total，再自己算
    if total == 0:
        total = inp + out
    return inp, out, total

# ---------- 新增：一次性扫描 figma url----------
def get_figma_urls(log_path: str) -> str:
    """
    从日志文件里提取所有 Figma URL，返回列表（按出现顺序，可重复）。
    如果文件不存在或没匹配到，返回空列表。
    """
    if not log_path or not os.path.isfile(log_path):
        return []
    #  分组 1 即为 URL
    pattern = re.compile(r"Received Figma URL:\s*(\S+)")
    with open(log_path, encoding="utf-8") as fh:
        for line in fh:
            m = pattern.search(line)
            if m:
                return m.group(1)


def format_tokens(inp, out, total):
    return f"input: {inp}\n, out: {out}\n,toatl: {total}\n"


def format_duration(sec):
    sec = int(sec)
    if sec < 60:
        return f"{sec}秒"
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    return f"{h}小时{m}分{s}秒" if h else f"{m}分{s}秒"

def find_images(fpath):
    img_dir = os.path.join(fpath, "app/src/test/snapshots/images/")
    if not os.path.exists(img_dir):
        return None, None
    a = b = None
    for f in os.listdir(img_dir):
        if f.startswith("com.example.myapplication"):
            a = os.path.join(img_dir, f)
        elif f.startswith("figma_screenshot"):
            b = os.path.join(img_dir, f)
        if a and b:
            break
    return a, b

# ---------- 新增：调用 compare.py 拿到 ratio & score ----------
import subprocess, os, sys, shlex

COMPARE_PY = os.path.join(os.path.dirname(os.path.abspath(__file__)), "image_compare.py")

def compare_images(design: str, actual: str):
    """
    返回 (ratio, score) 字符串；失败返回 ('N/A','N/A')
    """
    if not os.path.isfile(design) or not os.path.isfile(actual):
        print(f"[compare] 图片不存在：{design} 或 {actual}")
        return "N/A", "N/A"

    cmd = [sys.executable, COMPARE_PY,
           design, actual, "--size", f"{image_width}x{image_height}"]   # 统一分辨率
    try:
        # 捕获 stdout+stderr，方便调试
        completed = subprocess.run(
            cmd, capture_output=True, text=True, check=True)
        ratio = score = "N/A"
        for line in completed.stdout.splitlines():
            if line.startswith("conv-diff="):
                ratio = line.split("=", 1)[1].strip()
            elif line.startswith("SSIM="):
                score = line.split("=", 1)[1].strip()
        print(f"[compare] ratio={ratio}, score={score}")
        return ratio, score

    except subprocess.CalledProcessError as e:
        # exit 1/2 都会走到这里
        print(f"[compare] 失败！exit={e.returncode}")
        print(f"cmd : {' '.join(shlex.quote(c) for c in cmd)}")
        print(f"stdout: {e.stdout}")
        print(f"stderr: {e.stderr}")
        return "N/A", "N/A"
    except Exception as e:
        print(f"[compare] 其他异常: {e}")
        return "N/A", "N/A"


# ---------- 主函数 ----------
def main():
    wb = Workbook()
    ws = wb.active
    ws.title = datetime.now().strftime('%Y%m%d-%H')

    folders = sorted(
        [f for f in os.listdir(base_dir)
         if os.path.isdir(os.path.join(base_dir, f)) and is_valid_folder(f)],
        key=str.lower
    )
    total = 0
    for f in folders:
        log = find_log_file(os.path.join(base_dir, f), f)
        total += calc_duration(log) if log else 0
    avg = total / len(folders) if folders else 0
    inp_sum = out_sum = tot_sum = 0
    # 1. 表头追加两列
    ws.append(["Folder Name", "D2C Image", "Figma Screenshot",
               f"Duration (avg {format_duration(avg)})", f"token (avg{format_tokens(inp_sum, out_sum, tot_sum)})",
               "Conv-Diff(<0.002)", "SSIM(>0.98)", "Greeting.kt 行数"])

    ws.column_dimensions['E'].alignment = Alignment(wrap_text=True, vertical='top')
    # 2. 列宽 +2
    for col, w in enumerate([30, 45, 45, 20, 20, 15, 15, 15], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    row = 2
    data_collect = {"duration": [],
                    "input_token": [],
                    "output_token": [],
                    "total_token": [],
                    "ratio": [],
                    "score": [],
                    "greeting_lines": []}
    for folder in folders:
        fpath = os.path.join(base_dir, folder)
        log_file = find_log_file(fpath, folder)
        dur = calc_duration(log_file) if log_file else 0
        figma_url = get_figma_urls(log_file)
        inp, out, tot = calc_tokens(log_file)
        img1, img2 = find_images(fpath)
        data_collect["duration"].append(dur)
        data_collect["input_token"].append(inp)
        data_collect["output_token"].append(out)
        data_collect["total_token"].append(tot)
        if img1 and img2:
            ws.cell(row=row, column=1, value=f"{folder}\n\n{figma_url}")
            # 插图
            for col, ip in enumerate([img1, img2], 2):
                img = XLImage(ip)
                img.width, img.height = image_width, image_height
                ws.add_image(img, f"{chr(64 + col)}{row}")
            ws.cell(row=row, column=4, value=format_duration(dur))
            ws.cell(row=row, column=5, value=format_tokens(inp, out, tot))               

            # 3. 新增：ratio & score
            ratio, score = map(float, compare_images(img2, img1))  # 参数顺序按需要调
            ws.cell(row=row, column=6, value=f"{ratio:.3%}")
            ws.cell(row=row, column=7, value=score)
            data_collect['ratio'].append(ratio)
            data_collect["score"].append(score)
            ws.row_dimensions[row].height = 600
            greeting_lines = find_greeting_lines(fpath)
            data_collect["greeting_lines"].append(greeting_lines)
            ws.cell(row=row, column=8, value=greeting_lines)

            row += 1
            ws.cell(row=row - 1, column=5).alignment = Alignment(wrap_text=True, vertical='bottom')
            ws.cell(row=row - 1, column=1).alignment = Alignment(wrap_text=True, vertical='bottom')
            ws.cell(row=row - 1, column=8).alignment = Alignment(wrap_text=True, vertical='bottom')
            
    for i in range(4, 8):
        ws.cell(row=row, column=i).alignment = Alignment(wrap_text=True, vertical='bottom')
    ws.cell(row, column=4, value=f"max:{format_duration(max(data_collect['duration']))}\n"
                                 f"min:{format_duration(min(data_collect['duration']))}\n"
                                 f"avg:{format_duration(sum(data_collect['duration'])/len(data_collect['duration']))}\n")
    ws.cell(row, column=5, value=f"input token max:{max(data_collect['input_token'])}\n"
                                 f"min:{min(data_collect['input_token'])}\n"
                                 f"avg:{sum(data_collect['input_token'])/len(data_collect['input_token'])}\n"
                                 f"output token max:{(max(data_collect['output_token']))}\n"
                                 f"min:{(min(data_collect['output_token']))}\n"
                                 f"avg:{sum(data_collect['output_token'])/len(data_collect['output_token'])}\n"
                                 f"total token max:{max(data_collect['total_token'])}\n"
                                 f"min:{(min(data_collect['total_token']))}\n"
                                 f"avg:{sum(data_collect['total_token'])/len(data_collect['total_token'])}\n")
    ws.cell(row, column=6, value=f"max:{max(data_collect['ratio']):.3%}\n"
                                 f"min:{min(data_collect['ratio']):.3%}\n"
                                 f"avg:{sum(data_collect['ratio'])/len(data_collect['ratio']):.3%}\n")
    ws.cell(row, column=7, value=f"max:{max(data_collect['score'])}\n"
                                 f"min:{min(data_collect['score'])}\n"
                                 f"avg:{sum(data_collect['score'])/len(data_collect['score'])}\n")
    ws.cell(row, column=8,
                           value=f"max:{max(data_collect['greeting_lines'])}\n"
                                 f"min:{min(data_collect['greeting_lines'])}\n"
                                f"avg:{sum(data_collect['greeting_lines'])/len(data_collect['greeting_lines']):.1f}\n")

    header_row = 1
    tok_cell = ws.cell(row=header_row, column=5, value="Token")
    tok_cell.alignment = Alignment(wrap_text=True)
    out_xlsx = f"{base_dir}/d2c_{datetime.now().strftime('%Y%m%d-%H')}_report.xlsx"
    wb.save(out_xlsx)
    print(f"报告已保存 为 {out_xlsx}，平均时间：{format_duration(avg)}")

if __name__ == "__main__":
    import sys
    main()
